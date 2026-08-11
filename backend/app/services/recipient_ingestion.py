"""Converts uploaded tabular recipient data (CSV or XLSX) into a normalized,
validated recipient dataset.

Provider-agnostic on purpose — this module knows nothing about Gmail or
email sending, it only turns untrusted bytes into rows. Any future
bulk-recipient action (a second email provider, SMS) reuses this without
touching provider code, and the Gmail provider never sees raw file bytes.

Per-row problems (bad email, empty row, duplicate) are collected as
``RecipientRowIssue`` entries rather than raised — one bad row must not
discard the good ones. Problems with the file itself (wrong type, too big,
unreadable, no header) raise ``RecipientFileError`` since there is nothing
partial to salvage.
"""

from __future__ import annotations

import csv
import io
from dataclasses import dataclass, field

from email_validator import EmailNotValidError, validate_email
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

_MAX_FILE_SIZE_BYTES = 5 * 1024 * 1024  # 5 MB
_MAX_ROWS = 5000
_EMAIL_COLUMN_CANDIDATES = ("email", "email address", "e-mail")


class RecipientFileError(Exception):
    """The file itself can't be processed — wrong type, too large, corrupt,
    or structurally missing what's needed (header row, email column).
    """


@dataclass(frozen=True)
class RecipientRowIssue:
    row_number: int  # 1-based, header row excluded
    reason: str
    email: str | None = None


@dataclass(frozen=True)
class NormalizedRecipient:
    email: str
    variables: dict[str, str]


@dataclass(frozen=True)
class RecipientParseResult:
    recipients: list[NormalizedRecipient]
    issues: list[RecipientRowIssue] = field(default_factory=list)
    total_rows: int = 0


def parse_recipient_file(*, filename: str, content: bytes) -> RecipientParseResult:
    if len(content) == 0:
        raise RecipientFileError("Uploaded file is empty.")
    if len(content) > _MAX_FILE_SIZE_BYTES:
        limit_mb = _MAX_FILE_SIZE_BYTES // (1024 * 1024)
        raise RecipientFileError(f"Uploaded file exceeds the {limit_mb} MB limit.")

    lowered = filename.lower()
    if lowered.endswith(".csv"):
        rows = _parse_csv(content)
    elif lowered.endswith(".xlsx"):
        rows = _parse_xlsx(content)
    else:
        raise RecipientFileError("Unsupported file type — upload a .csv or .xlsx file.")

    if len(rows) > _MAX_ROWS:
        raise RecipientFileError(f"File has more than {_MAX_ROWS} rows.")

    return _normalize(rows)


def _parse_csv(content: bytes) -> list[dict[str, str]]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise RecipientFileError("CSV file is not valid UTF-8 text.") from exc
    reader = csv.DictReader(io.StringIO(text))
    if reader.fieldnames is None:
        raise RecipientFileError("CSV file has no header row.")
    return [dict(row) for row in reader]


def _parse_xlsx(content: bytes) -> list[dict[str, str]]:
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except (InvalidFileException, KeyError, OSError) as exc:
        raise RecipientFileError("Could not read file as an XLSX workbook.") from exc

    sheet = workbook.active
    if sheet is None:
        raise RecipientFileError("XLSX workbook has no sheets.")

    rows_iter = sheet.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if header is None:
        raise RecipientFileError("XLSX sheet has no header row.")

    headers = [str(cell).strip() if cell is not None else "" for cell in header]
    rows: list[dict[str, str]] = []
    for raw_row in rows_iter:
        row = {
            headers[i]: ("" if raw_row[i] is None else str(raw_row[i]))
            for i in range(len(headers))
            if headers[i] and i < len(raw_row)
        }
        rows.append(row)
    return rows


def _normalize(rows: list[dict[str, str]]) -> RecipientParseResult:
    if not rows:
        raise RecipientFileError("File has no data rows.")

    email_column = _find_email_column(rows[0])
    if email_column is None:
        raise RecipientFileError(
            "No email column found — include a column named 'email', 'Email Address', or similar."
        )

    recipients: list[NormalizedRecipient] = []
    issues: list[RecipientRowIssue] = []
    seen_emails: set[str] = set()

    for index, row in enumerate(rows, start=1):
        variables = {
            key.strip(): (value or "").strip()
            for key, value in row.items()
            if key and key.strip() and key.strip() != email_column
        }
        raw_email = (row.get(email_column) or "").strip()

        if not raw_email and not any(variables.values()):
            issues.append(RecipientRowIssue(row_number=index, reason="empty row"))
            continue

        if not raw_email:
            issues.append(RecipientRowIssue(row_number=index, reason="missing email"))
            continue

        try:
            validated = validate_email(raw_email, check_deliverability=False)
        except EmailNotValidError as exc:
            issues.append(
                RecipientRowIssue(row_number=index, reason=f"invalid email: {exc}", email=raw_email)
            )
            continue

        normalized_email = validated.normalized.lower()
        if normalized_email in seen_emails:
            issues.append(
                RecipientRowIssue(
                    row_number=index, reason="duplicate email", email=normalized_email
                )
            )
            continue

        seen_emails.add(normalized_email)
        recipients.append(NormalizedRecipient(email=normalized_email, variables=variables))

    return RecipientParseResult(recipients=recipients, issues=issues, total_rows=len(rows))


def _find_email_column(sample_row: dict[str, str]) -> str | None:
    for key in sample_row:
        if key and key.strip().lower() in _EMAIL_COLUMN_CANDIDATES:
            return key
    return None
